import tkinter as tk
from tkinter import filedialog, messagebox
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from openpyxl import load_workbook
import os


def formatear_numero(numero):
    numero_formateado = "{:,}".format(numero).replace(",", "X").replace(".", ",").replace("X", ".")
    return numero_formateado

def generar_desde_excel(ruta_excel):
    try:
        wb = load_workbook(ruta_excel)
        ws = wb.active

        for fila in ws.iter_rows(min_row=2, values_only=True):
            nombre, cedula, valor, concepto, rango, firma = fila

            if not nombre:
                continue  # evita filas vacías

            # Limpieza de datos
            valor_str = formatear_numero(valor)
            cedula_str = formatear_numero(cedula)

            # Validar firma
            if firma and not os.path.exists(firma):
                #Si no encuentra la fimra, simplmenete es null
                firma = None

            generar_pdf(
                str(nombre),
                cedula_str,
                valor_str,
                str(concepto),
                str(rango),
                firma
            )

        messagebox.showinfo("Listo", "Todos los PDFs fueron generados correctamente")

    except Exception as e:
        messagebox.showerror("Error", str(e))

def fecha_en_espanol():
    meses = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO",
        4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE",
        10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }

    hoy = datetime.now()
    return f"{hoy.day} DE {meses[hoy.month]} DE {hoy.year}"

# ---------------------------
# NUMERO A LETRAS (MEJORADO)
# ---------------------------
def numero_a_letras(num):
    unidades = (
        '', 'uno', 'dos', 'tres', 'cuatro', 'cinco', 'seis',
        'siete', 'ocho', 'nueve'
    )

    especiales = (
        'diez', 'once', 'doce', 'trece', 'catorce', 'quince',
        'dieciseis', 'diecisiete', 'dieciocho', 'diecinueve'
    )

    decenas = (
        '', '', 'veinte', 'treinta', 'cuarenta',
        'cincuenta', 'sesenta', 'setenta', 'ochenta', 'noventa'
    )

    centenas = (
        '', 'ciento', 'doscientos', 'trescientos', 'cuatrocientos',
        'quinientos', 'seiscientos', 'setecientos', 'ochocientos', 'novecientos'
    )

    def convertir(n):
        if n < 10:
            return unidades[n]
        elif n < 20:
            return especiales[n-10]
        elif n < 30:
            return "veinti" + unidades[n-20]
        elif n < 100:
            d, u = divmod(n, 10)
            return decenas[d] + (" y " + unidades[u] if u else "")
        elif n == 100:
            return "cien"
        elif n < 1000:
            c, r = divmod(n, 100)
            return centenas[c] + (" " + convertir(r) if r else "")
        elif n < 1000000:
            m, r = divmod(n, 1000)
            texto = "mil" if m == 1 else convertir(m) + " mil"
            return texto + (" " + convertir(r) if r else "")
        elif n < 1000000000:
            m, r = divmod(n, 1000000)
            if m == 1:
                texto = "un millon"
            else:
                texto = convertir(m) + " millones"
            return texto + (" " + convertir(r) if r else "")
        else:
            return str(n)

    texto = convertir(num).upper()

    # Ajuste gramatical
    if "MILLON" in texto or "MILLONES" in texto:
        texto += " DE PESOS M/CTE"
    else:
        texto += " PESOS M/CTE"

    return texto


# ---------------------------
# GENERAR PDF
# ---------------------------
def generar_pdf(nombre, cedula, precio,concepto, rango, firma_path):

    pdfmetrics.registerFont(TTFont("TimesNewRoman", "C:/Windows/Fonts/times.ttf"))
    pdfmetrics.registerFont(TTFont("TimesNewRoman-Bold", "C:/Windows/Fonts/timesbd.ttf"))
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.alignment = 1  # 0=izq, 1=centrado, 2=derecha
    style.fontName = "Times-Roman"
    style.fontSize = 11

    fecha = fecha_en_espanol()

    valor_num = int(precio.replace(".", "").replace(",", ""))
    letras = numero_a_letras(valor_num)

    archivo = f"CUENTA DE COBRO {nombre.upper()} {rango}.pdf"
    c = canvas.Canvas(archivo, pagesize=letter)
    width, height = letter
    y = height - 50
    # Fecha (izquierda)
    c.setFont("Times-Roman", 11)
    c.drawString(50, y, f"ACACIAS {fecha}")

    y -= 100
    # Encabezado centrado
    c.setFont("Times-Bold", 12)
    c.drawCentredString(width/2, y, "LINEAS JAVARTURS DEL LLANO SAS")

    y -= 20
    c.setFont("Times-Roman", 11)
    c.drawCentredString(width/2, y, "NIT: 830.085.825-5")

    y -= 70

    c.drawCentredString(width/2, y, "DEBE A:")
    y -= 25

    c.setFont("Times-Bold", 11)
    c.drawCentredString(width/2, y, nombre.upper())
    y -= 20

    c.setFont("Times-Roman", 11)
    c.drawCentredString(width/2, y, f"C.C. {cedula}")
    y -= 50

    texto_letras = Paragraph(f"<b>LA SUMA DE: ${precio}</b> ({letras})", style)

    w, h = texto_letras.wrap(width - 250, height)
    texto_letras.drawOn(c, 125, y-h+10)

    y -= h + 20

    

    texto_concepto = Paragraph(f"<b>POR CONCEPTO DE:</b> {concepto} {rango}", style)

    w, h = texto_concepto.wrap(width - 250, height)
    texto_concepto.drawOn(c, 125, y-h+10)

    y -= h + 90

    c.drawString(50, y, "ATENTAMENTE:")
    y -= 80

    # Línea firma (izquierda)
    x_firma = 50
    c.line(x_firma, y, x_firma + 250, y)

    # Firma EXACTAMENTE sobre la línea
    if firma_path:
        try:
            c.drawImage(
                firma_path,
                x_firma,
                y - 15,   # ajuste fino
                width=200,
                height=50,
                mask='auto'
            )
        except:
            pass

    y -= 30
    c.drawString(x_firma, y, f"NOMBRE: {nombre.upper()}")
    y -= 20
    c.drawString(x_firma, y, f"C.C.: {cedula}")

    c.save()
    # messagebox.showinfo("Listo", f"PDF generado:\n{archivo}")


# ---------------------------
# INTERFAZ
# ---------------------------
def cargar_firma():
    ruta = filedialog.askopenfilename(filetypes=[("Imagen", "*.png *.jpg *.jpeg")])
    entry_firma.delete(0, tk.END)
    entry_firma.insert(0, ruta)

def cargar_excel():
    ruta = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if ruta:
        generar_desde_excel(ruta)

def generar():
    generar_pdf(
        entry_nombre.get(),
        entry_cedula.get(),
        entry_precio.get(),
        entry_concepto.get(),
        entry_rango.get(),
        entry_firma.get()
    )


root = tk.Tk()
root.title("Cuenta de Cobro PRO")

labels = ["Nombre", "Cédula", "Valor ($)", "Concepto", "Rango trabajado"]
entries = []

for i, text in enumerate(labels):
    tk.Label(root, text=text).grid(row=i, column=0)
    e = tk.Entry(root, width=40)
    e.grid(row=i, column=1)
    entries.append(e)

entry_nombre, entry_cedula, entry_precio, entry_concepto, entry_rango = entries

tk.Label(root, text="Firma").grid(row=5, column=0)
entry_firma = tk.Entry(root, width=30)
entry_firma.grid(row=5, column=1)

tk.Button(root, text="Cargar", command=cargar_firma).grid(row=6, column=2)
tk.Button(root, text="Generar PDF", command=generar).grid(row=6, column=1, pady=10)
tk.Button(root, text="Cargar Excel", command=cargar_excel).grid(row=7, column=1, pady=10)
root.mainloop()